import pathlib
import openpyxl
import csv
"""
EXCWLデータ練習
"""


lwb = openpyxl.Workbook() 

lsh = lwb.active

list_row = 1
path = pathlib.Path("./")
for pass_obj in path.iterdir():
    if pass_obj.match("*.xlsx"):
        wb = openpyxl.load_workbook(pass_obj)
        for sh in wb:
            for dt_row in range(9,19):
                if sh.cell(dt_row, 2).value != None:
                    lsh.cell(list_row,1).value = sh.cell(2,7).value
                    list_row += 1

with open("./saleslist.csv","w",encoding="utf_8_sig") as fp:
    writer = csv.writer(fp,lineterminator="\n")
    for row in lsh.rows:
        writer.writerow([col.value for col in row])

